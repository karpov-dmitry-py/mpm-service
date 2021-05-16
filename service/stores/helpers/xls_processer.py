import os.path
import shutil
import uuid
from datetime import datetime

import openpyxl
from django.http import HttpResponse

from .common import _log
from .common import _err
from .common import _exc
from ..models import GoodsBrand
from ..models import GoodsCategory
from ..models import Good


class ExcelProcesser:
    BASE_DIR = '/tmp/mpm-goods-batch-processing'

    def __init__(self, check_tmp_dir=True):
        if check_tmp_dir:
            self._check_tmp_dir()

    def get_save_dir(self):
        return self.BASE_DIR

    @staticmethod
    def _now_as_str():
        return datetime.now().strftime('%d%m%Y_%H%M%S')

    def _check_tmp_dir(self):
        _dir = self.BASE_DIR
        if not os.path.exists(_dir):
            os.makedirs(_dir, exist_ok=True)

    def export_goods(self, queryset):
        data_first_row = 3
        dirname = os.path.dirname(__file__)
        template_file = 'templates/goods/goods_upload_sample.xlsx'
        template_path = os.path.join(dirname, template_file)

        filename = f'goods_export_{self._now_as_str()}.xlsx'
        _dir = self._new_dir()
        path = os.path.join(self.BASE_DIR, _dir)
        os.makedirs(path, exist_ok=True)
        file_full_path = os.path.join(path, filename)

        items = []
        for row in queryset:
            item = {
                'sku': row.sku,
                'article': row.article,
                'name': row.name,
                'brand': str(row.brand) if row.brand else None,
                'category': row.category.name if row.category else None,
                'barcode': row.barcode,
                'pack_length': row.pack_length,
                'pack_width': row.pack_width,
                'pack_height': row.pack_height,
                'weight': row.weight,
            }
            items.append(item)

        fields = [
            {
                'name': 'Комментарий',
                'field': None,
                'description': 'Комментарий для результатов загрузки',
                'required': False,
            },
            {
                'name': 'Ваш SKU',
                'field': 'sku',
                'description': 'Уникальный идентификатор товара в учетной системе',
                'required': True,
            },
            {
                'name': 'Артикул производителя',
                'field': 'article',
                'description': '',
                'required': False,
            },
            {
                'name': 'Наименование товара',
                'field': 'name',
                'description': '',
                'required': True,
            },
            {
                'name': 'Бренд',
                'field': 'brand',
                'description': 'Заполняется согласно справочнику "Бренды"',
                'required': False,
            },
            {
                'name': 'Категория товара',
                'field': 'category',
                'description': 'Указывается конечная категория (нижнего уровня) согласно справочнику "Категории"',
                'required': False,
            },
            {
                'name': 'Штрихкод',
                'field': 'barcode',
                'description': '',
                'required': False,
            },
            {
                'name': 'Длина упаковки, см',
                'field': 'pack_length',
                'description': '',
                'required': False,
            },
            {
                'name': 'Ширина упаковки, см',
                'field': 'pack_width',
                'description': '',
                'required': False,
            },
            {
                'name': 'Высота упаковки, см',
                'field': 'pack_height',
                'description': '',
                'required': False,
            },
            {
                'name': 'Вес, кг',
                'field': 'weight',
                'description': '',
                'required': False,
            },
        ]

        wb = openpyxl.load_workbook(template_path)
        sheet = wb.active

        # write headers
        for row, item in enumerate(items, start=data_first_row):
            for col, field in enumerate(fields, start=1):
                cell = sheet.cell(row=row, column=col)
                if field_name := field.get('field'):
                    value = item.get(field_name)
                    if value is not None:
                        cell.value = value

        wb.save(file_full_path)
        file_path = f'{_dir}&&{filename}'
        return self.get_file_content(file_path)

    @staticmethod
    def _new_dir():
        return str(uuid.uuid4())

    @staticmethod
    def _val_to_float(val):
        if not val:
            return {'float': None}
        if isinstance(val, float):
            return {'float': val}
        if isinstance(val, str):
            patterns = (',',)
            for item in patterns:
                val = val.replace(item, '.')
        try:
            _float = float(val)
            return {'float': _float}
        except (TypeError, Exception) as err:
            err_msg = f'Failed to convert value {str(val)} to float: {_exc(err)}'
            _err(err_msg)
            err_msg = f' Не удалось преобразовать значение {str(val)} в десятичное число: {_exc(err)}'
            return {'error': err_msg}

    def batch_goods_upload(self, raw, file_extension, user):
        # save raw to file
        filename = f'goods_upload_{self._now_as_str()}{file_extension}'
        save_dir = self._new_dir()
        dir_path = os.path.join(self.BASE_DIR, save_dir)
        os.makedirs(dir_path)
        full_path = os.path.join(dir_path, filename)
        with open(full_path, 'wb') as file:
            file.write(raw)

        # read xlsx
        xls = openpyxl.load_workbook(full_path)
        sheet = xls.active
        _rows = []
        cols = []
        row_num_data_start = 3
        for row_number, row in enumerate(sheet.iter_rows(), start=1):
            if row_number == 1:
                for cell_number, cell in enumerate(row, start=1):
                    cols.append(cell.value)
                continue
            elif row_number < row_num_data_start:
                continue
            _row = dict()
            for cell_number, cell in enumerate(row, start=1):
                _row[cols[cell_number - 1]] = cell.value
                _row['row_number'] = row_number
                _row['comment'] = ''
            _rows.append(_row)

        # read existing rows from db
        # noinspection PyUnresolvedReferences
        goods_set = Good.objects.filter(user=user)
        goods = {item.sku: item for item in goods_set}
        # noinspection PyUnresolvedReferences
        brand_set = GoodsBrand.objects.filter(user=user)
        brands = {item.name.lower().strip(): item for item in brand_set}
        # noinspection PyUnresolvedReferences
        categories_set = GoodsCategory.objects.filter(user=user)
        categories = {item.name.lower().strip(): item for item in categories_set}

        # process data
        stats = {
            'created_goods': set(),
            'skipped_goods': set(),
            'updated_goods': set(),
            'goods_db_error': set(),
        }

        empty_field_err = 'Ошибка. Не заполнено обязательное поле:'
        for row in _rows:
            cleaned_data = {'user': user}

            # sku
            sku_col_label = 'Ваш SKU'
            sku = str(row[sku_col_label]).strip()
            if not sku:
                row['comment'] = f'{empty_field_err} "{sku_col_label}"'
                continue
            cleaned_data['sku'] = sku

            # article
            article_col_label = 'Артикул производителя'
            article = row[article_col_label]
            cleaned_data['article'] = article

            # name
            name_col_label = 'Наименование товара'
            name = str(row[name_col_label]).strip()
            if not name:
                row['comment'] = f'{empty_field_err} "{name_col_label}"'
                continue
            cleaned_data['name'] = name

            # brand
            brand_col_label = 'Бренд'
            brand = row[brand_col_label]
            brand = brand or ''
            brand_ref = brands.get(brand.lower())
            if brand and not brand_ref:
                new_brand = GoodsBrand(name=brand, user=user)
                try:
                    new_brand.save()
                    brands[brand.lower()] = new_brand
                    brand_ref = new_brand
                    _log(f'A new brand created: "{brand}"')
                except Exception as err:
                    err_msg = f'Failed to create a new brand: "{brand}": {_exc(err)}'
                    _err(err_msg)
                    err_msg = f'Ошибка. Не удалось создать новый бренд "{brand}": {_exc(err)}'
                    row['comment'] = err_msg
            cleaned_data['brand'] = brand_ref

            # category
            category_col_label = 'Категория товара'
            category = row[category_col_label]
            category = category or ''
            category_ref = categories.get(category.lower())
            cleaned_data['category'] = category_ref

            # barcode
            barcode_col_label = 'Штрихкод'
            barcode = row[barcode_col_label]
            cleaned_data['barcode'] = barcode

            # pack_length
            pack_length_col_label = 'Длина упаковки, см'
            pack_length = row[pack_length_col_label] or None
            result = self._val_to_float(pack_length)
            if error := result.get('error'):
                row['comment'] += f'{error}'
                continue
            cleaned_data['pack_length'] = result['float']

            # pack_width
            pack_width_col_label = 'Ширина упаковки, см'
            pack_width = row[pack_width_col_label] or None
            result = self._val_to_float(pack_width)
            if error := result.get('error'):
                row['comment'] += f'{error}'
                continue
            cleaned_data['pack_width'] = result['float']

            # pack_height
            pack_height_col_label = 'Высота упаковки, см'
            pack_height = row[pack_height_col_label] or None
            result = self._val_to_float(pack_height)
            if error := result.get('error'):
                row['comment'] += f'{error}'
                continue
            cleaned_data['pack_height'] = result['float']

            # weight
            weight_col_label = 'Вес, кг'
            weight = row[weight_col_label] or None
            result = self._val_to_float(weight)
            if error := result.get('error'):
                row['comment'] += f'{error}'
                continue
            cleaned_data['weight'] = result['float']

            # existing good
            if sku in goods:
                existing_good = goods[sku]
                update_required = False
                for key, value in cleaned_data.items():
                    if getattr(existing_good, key) != value:
                        setattr(existing_good, key, value)
                        update_required = True
                if update_required:
                    # update_required {
                    try:
                        existing_good.save()
                        _log(f'An existing good updated: "{existing_good}"')
                        row['comment'] = f'Обновлен в БД. {row["comment"]}'
                        stats['updated_goods'].add(existing_good.sku)
                    except Exception as err:
                        err_msg = f'Failed to create a new good: "({sku}) {name}": {_exc(err)}'
                        _err(err_msg)
                        err_msg = f'Ошибка. Не удалось обновить существующий товар: {_exc(err)}'
                        row['comment'] = f'{err_msg} {row["comment"]}'
                        stats['goods_db_error'].add(existing_good.sku)

                    # update_required }
                else:
                    _log(f'Skipped an existing good (no db update is required): ({sku}) "{name}"')
                    row['comment'] = f'Пропущен (данные в БД актуальны). {row["comment"]}'
                    stats['skipped_goods'].add(existing_good.sku)
            # new good
            else:
                new_good = Good(**cleaned_data)
                try:
                    new_good.save()
                    goods[sku] = new_good
                    _log(f'A new good created: "{new_good}"')
                    row['comment'] = f'Добавлен в БД. {row["comment"]}'
                    stats['created_goods'].add(new_good.sku)
                except Exception as err:
                    err_msg = f'Failed to create a new good: "{new_good}": {_exc(err)}'
                    _err(err_msg)
                    err_msg = f'Ошибка. Не удалось создать новый товар: {_exc(err)}'
                    row['comment'] = f'{err_msg} {row["comment"]}'
                    stats['goods_db_error'].add(sku)

        for row in _rows:
            cell = f'A{row["row_number"]}'
            sheet[cell].value = row['comment']

        # append stats to xls
        stats_sheet = xls.create_sheet(title="Статистика загрузки")
        row_number = 0
        for key, value in stats.items():
            row_number += 1
            stats_sheet[f'A{row_number}'] = f'{key}:'
            stats_sheet[f'B{row_number}'] = len(value)

        xls.save(filename=full_path)
        xls.close()
        return save_dir, filename

    def get_file_content(self, path):
        try:
            parts = path.split('&&')
            _dir, filename = parts[0], parts[1]
        except (IndexError, Exception):
            err_msg = f'Не корректные данные файла: {path}'
            return None, err_msg
        relative_path_to_file = os.path.join(_dir, filename)
        fullpath = os.path.join(self.BASE_DIR, relative_path_to_file)
        if not os.path.isfile(fullpath):
            err_msg = f'Не найден файл: {relative_path_to_file}'
            return None, err_msg
        with open(fullpath, 'rb') as file:
            raw = file.read()
        filename = os.path.basename(fullpath)
        shutil.rmtree(_dir, ignore_errors=True)
        response = HttpResponse(raw, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response, None
